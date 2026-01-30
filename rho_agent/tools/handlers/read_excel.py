"""Read Excel file contents handler."""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..base import ToolHandler, ToolInvocation, ToolOutput

# Max rows to return by default
DEFAULT_MAX_ROWS = 500


def is_row_visible(ws: Worksheet, row: int) -> bool:
    """Check if a row is visible (not hidden or in a collapsed group)."""
    rd = ws.row_dimensions.get(row)
    if rd is None:
        return True
    if rd.hidden:
        return False
    # Rows with outlineLevel > 0 that are hidden are collapsed
    if rd.outlineLevel and rd.outlineLevel > 0 and rd.hidden:
        return False
    return True


def is_col_visible(ws: Worksheet, col: int) -> bool:
    """Check if a column is visible (not hidden or in a collapsed group)."""
    col_letter = get_column_letter(col)
    cd = ws.column_dimensions.get(col_letter)
    if cd is None:
        return True
    if cd.hidden:
        return False
    if cd.outlineLevel and cd.outlineLevel > 0 and cd.hidden:
        return False
    return True


def format_cell_value(value: Any) -> str:
    """Format a cell value for display."""
    if value is None:
        return ""
    if isinstance(value, float):
        # Avoid excessive decimal places
        if value == int(value):
            return str(int(value))
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value)


class ReadExcelHandler(ToolHandler):
    """Read and inspect Excel files (.xlsx, .xls)."""

    @property
    def name(self) -> str:
        return "read_excel"

    @property
    def description(self) -> str:
        return (
            "Read and inspect Excel files. Supports three actions:\n"
            "- list_sheets: List all sheet names in the workbook\n"
            "- read_sheet: Read contents of a specific sheet (respects hidden rows/columns by default)\n"
            "- get_info: Get workbook metadata (sheet count, dimensions per sheet)\n"
            "By default, hidden rows and columns are excluded. Use show_hidden=true to include them."
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "Absolute path to the Excel file (.xlsx or .xls)",
                },
                "action": {
                    "type": "string",
                    "enum": ["list_sheets", "read_sheet", "get_info"],
                    "description": "Action to perform",
                },
                "sheet": {
                    "type": "string",
                    "description": "Sheet name for read_sheet action. If omitted, reads the first sheet.",
                },
                "start_row": {
                    "type": "integer",
                    "description": "First row to read (1-indexed). Defaults to 1.",
                },
                "end_row": {
                    "type": "integer",
                    "description": "Last row to read (1-indexed). Defaults to start_row + 500.",
                },
                "show_hidden": {
                    "type": "boolean",
                    "description": "Include hidden rows and columns. Defaults to false.",
                },
            },
            "required": ["path", "action"],
        }

    async def handle(self, invocation: ToolInvocation) -> ToolOutput:
        path_str = invocation.arguments.get("path", "")
        action = invocation.arguments.get("action", "")

        if not path_str:
            return ToolOutput(content="No path provided", success=False)

        if not action:
            return ToolOutput(content="No action provided", success=False)

        path = Path(path_str).expanduser().resolve()

        if not path.exists():
            return ToolOutput(content=f"File not found: {path}", success=False)

        if not path.is_file():
            return ToolOutput(content=f"Not a file: {path}", success=False)

        try:
            # Load workbook (data_only to get values not formulas)
            # Note: can't use read_only=True because we need row/column dimensions for visibility
            wb = load_workbook(path, data_only=True)

            if action == "list_sheets":
                return self._list_sheets(wb)
            elif action == "get_info":
                return self._get_info(wb, path)
            elif action == "read_sheet":
                return self._read_sheet(wb, invocation.arguments)
            else:
                return ToolOutput(content=f"Unknown action: {action}", success=False)

        except Exception as e:
            return ToolOutput(content=f"Error reading Excel file: {e}", success=False)

    def _list_sheets(self, wb) -> ToolOutput:
        """List all sheet names."""
        sheets = wb.sheetnames
        content = f"Sheets ({len(sheets)}):\n"
        for i, name in enumerate(sheets, 1):
            content += f"  {i}. {name}\n"
        return ToolOutput(
            content=content.strip(),
            success=True,
            metadata={"sheet_count": len(sheets), "sheets": sheets},
        )

    def _get_info(self, wb, path: Path) -> ToolOutput:
        """Get workbook metadata."""
        lines = [f"File: {path.name}", f"Sheets: {len(wb.sheetnames)}", ""]

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"  {sheet_name}:")
            lines.append(f"    Dimensions: {ws.dimensions or 'empty'}")
            if ws.max_row and ws.max_column:
                lines.append(f"    Rows: {ws.max_row}, Columns: {ws.max_column}")

        return ToolOutput(
            content="\n".join(lines),
            success=True,
            metadata={"sheets": wb.sheetnames},
        )

    def _read_sheet(self, wb, args: dict[str, Any]) -> ToolOutput:
        """Read contents of a sheet."""
        sheet_name = args.get("sheet")
        start_row = args.get("start_row", 1)
        end_row = args.get("end_row")
        show_hidden = args.get("show_hidden", False)

        # Get the sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return ToolOutput(
                    content=f"Sheet not found: {sheet_name}. Available: {', '.join(wb.sheetnames)}",
                    success=False,
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
            sheet_name = ws.title

        # Handle empty sheets
        if not ws.max_row or not ws.max_column:
            return ToolOutput(
                content=f"Sheet '{sheet_name}' is empty",
                success=True,
                metadata={"sheet": sheet_name, "rows": 0, "columns": 0},
            )

        # Determine row range
        start_row = max(1, start_row)
        if end_row is None:
            end_row = min(start_row + DEFAULT_MAX_ROWS - 1, ws.max_row)
        else:
            end_row = min(end_row, ws.max_row)

        # Determine visible columns
        visible_cols = []
        for col in range(1, ws.max_column + 1):
            if show_hidden or is_col_visible(ws, col):
                visible_cols.append(col)

        if not visible_cols:
            return ToolOutput(
                content=f"Sheet '{sheet_name}' has no visible columns",
                success=True,
                metadata={"sheet": sheet_name},
            )

        # Build output
        lines = []
        hidden_row_count = 0
        visible_row_count = 0

        for row in range(start_row, end_row + 1):
            row_visible = show_hidden or is_row_visible(ws, row)

            if not row_visible:
                hidden_row_count += 1
                continue

            visible_row_count += 1
            row_values = []
            for col in visible_cols:
                cell = ws.cell(row=row, column=col)
                row_values.append(format_cell_value(cell.value))

            # Mark hidden rows when show_hidden is True
            prefix = "[H] " if show_hidden and not is_row_visible(ws, row) else ""
            lines.append(prefix + "\t".join(row_values))

        # Build header with column letters
        col_headers = [get_column_letter(c) for c in visible_cols]
        header_line = "\t".join(col_headers)

        # Summary
        total_rows = ws.max_row
        total_cols = ws.max_column
        visible_cols_count = len(visible_cols)
        hidden_cols_count = total_cols - visible_cols_count

        summary = f"Sheet: {sheet_name} ({total_rows} rows x {total_cols} cols"
        if not show_hidden and (hidden_row_count > 0 or hidden_cols_count > 0):
            summary += f", showing {visible_row_count} rows x {visible_cols_count} cols"
        summary += ")"

        content = summary + "\n" + header_line + "\n" + "\n".join(lines)

        # Add footer note about truncation/hidden
        notes = []
        if end_row < ws.max_row:
            notes.append(f"Showing rows {start_row}-{end_row} of {ws.max_row}")
        if not show_hidden and hidden_row_count > 0:
            notes.append(f"{hidden_row_count} rows hidden")
        if not show_hidden and hidden_cols_count > 0:
            notes.append(f"{hidden_cols_count} columns hidden")

        if notes:
            content += f"\n\n[{', '.join(notes)}]"

        return ToolOutput(
            content=content,
            success=True,
            metadata={
                "sheet": sheet_name,
                "start_row": start_row,
                "end_row": end_row,
                "total_rows": total_rows,
                "visible_rows": visible_row_count,
                "total_columns": total_cols,
                "visible_columns": visible_cols_count,
            },
        )
